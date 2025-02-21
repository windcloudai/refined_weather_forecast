from openpyxl.styles import PatternFill
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font


def get_colour(cell):
    color_hex = ''
    # 获取单元格填充样式
    if cell.fill.start_color.index:
        color_index = cell.fill.start_color.index
        # 将索引颜色转换为RGB（适用于索引色和主题色）
        if isinstance(color_index, int):
            rgb = cell.fill.start_color.tint
            color_hex = f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
        else:
            color_hex = color_index[2:]  # 去除ARGB前缀
    return color_hex

def get_color_mapping(ws_sheet2):
    """从Sheet2获取颜色与天气的映射关系"""
    color_map = {}
    # color_map['000000'] = '没有'
    for row in ws_sheet2.iter_rows(min_row=1):
        for cell in row:
            if get_colour(cell) == '000000':
                # color_map['000000'] = '没有'
                continue
            if cell.value is None:
                break
            color_hex = get_colour(cell)
            color_map[color_hex] = cell.value
    # print('天气集合color_map-->', color_map)
    return color_map

def process_weather_data(file_path):
    # 加载工作簿
    wb = load_workbook(file_path)
    # 处理Sheet2的颜色映射
    ws_sheet2 = wb["Sheet1"]  # 根据实际Sheet名称修改

    color_weather_map = get_color_mapping(ws_sheet2)
    # 处理Sheet1的数据
    ws_sheet1 = wb["Sheet"]  # 根据实际Sheet名称修改

    # ws_sheet1 = ws.active
    grouped_data = defaultdict(set)
    grouped_color = defaultdict(int)
    grouped_dayujiaxue_cell = defaultdict(list)
    grouped_daxue_cell = defaultdict(list)

    # 确定列位置（根据实际表头调整）
    line_col = 1  # 线路名列
    value_col = 4  # 48小时数据列
    color_48hour_set = set()

    for row in ws_sheet1.iter_rows(min_row=4):  # 从数据行开始
        line_name = row[line_col - 1].value  # 线路名
        if line_name is None:
            break
        cell_48hour = row[value_col - 1]  # 第4个单元格：48小时列
        # print(f'cell-->{cell_48hour}')
        # print(f'line_name-->{line_name},cell-->{cell_48hour.value}')
        # 获取列'48小时' 单元格的颜色
        color_48hour = get_colour(cell_48hour)
        color_48hour_set.add(color_48hour)
        # print(f'color_48hour-->{color_48hour}')
        # print(f'color_48hour_set-->{color_48hour_set}')
        grouped_data[line_name].add(color_weather_map[color_48hour])
        # print(f'grouped_data-->{grouped_data}')
        # 遍历当前行的每个单元格
        for cell in row:
            # 获取当前单元格的列索引
            col_index = cell.column
            # print('行号:',cell.row)
            # print(f'col_index-->{col_index}')
            # 获取第3行表头中对应列的单元格
            header_cell = ws_sheet1.cell(row=3, column=col_index)

            # 打印当前单元格的值和对应的表头值
            # print(f"单元格 {cell.coordinate} 的值: {cell.value}，属于第{col_index}列，属于第3行表头的列：{header_cell.value}")
            # print(f'grouped_data.values()-->{grouped_data.values()}')
            # 获取每个时间段的单元格颜色
            cell_color = get_colour(cell)
            if cell.column >= 5 and cell_color in color_48hour_set:
                # print('满足条件的cell：', cell)
                if grouped_color[line_name + '_mincol'] == 0:
                    grouped_color[line_name + '_mincol'] = cell.column
                    grouped_color[line_name + '_mincell'] = cell
                    grouped_color[line_name + '_minheadcellvalue'] = ws_sheet1.cell(row=3, column=cell.column).value
                elif cell.column < grouped_color[line_name + '_mincol']:
                    grouped_color[line_name + '_mincol'] = cell.column
                    grouped_color[line_name + '_mincell'] = cell
                    grouped_color[line_name + '_minheadcellvalue'] = ws_sheet1.cell(row=3, column=cell.column).value

                if grouped_color[line_name + '_maxcol'] == 0:
                    grouped_color[line_name + '_maxcol'] = cell.column
                    grouped_color[line_name + '_maxcell'] = cell
                    grouped_color[line_name + '_maxheadcellvalue'] = ws_sheet1.cell(row=3, column=cell.column).value
                elif cell.column > grouped_color[line_name + '_maxcol']:
                    grouped_color[line_name + '_maxcol'] = cell.column
                    grouped_color[line_name + '_maxcell'] = cell
                    grouped_color[line_name + '_maxheadcellvalue'] = ws_sheet1.cell(row=3, column=cell.column).value

                # print("每一个行，每一个单元格的grouped_color是：",  grouped_color)

            # 获取【其中】的文字，找“大雨夹雪”的红色单元格
            # 在一个线路里，查找红色单元格，获取红色单元格的最小和最大的列的值
            # 判断：第2行红色单元格的行 == 第1行红色单元格的行 + 1 and ，放到一个组别里，否则就是另一个组别
            # 找最大列，最小列，最大行、最小行
            # 大雨夹雪
            if cell.column >= 5 and cell_color == 'DE4D21':
                grouped_dayujiaxue_cell[line_name].append(cell)
            # print()

    # print("grouped_data是：",  grouped_data)
    # print("grouped_dayujiaxue_cell是：",  grouped_dayujiaxue_cell)
    # print("grouped_daxue_cell：",  grouped_daxue_cell)

    # 生成文字结果
    # 【总】的文字
    result = defaultdict(list)
    xueresults = ''
    for line_name, weathers in grouped_data.items():
        if weathers:
            # print(f'weathers-->{weathers}')
            weather_list = sorted(list(weathers))
            # print(f'weather_list-->{weather_list}')
            A_without_B = [x for x in weather_list if x.find('夹') == -1]
            if len(A_without_B) == 0:
                A_without_B = [item.replace('雨夹', '') for item in weathers]
            # print(f'A_without_B-->{A_without_B}')
            # print()
            if len(A_without_B) == 3:
                xueresults = '小到大雪'
            else:
                xueresults = '、'.join(A_without_B)
                xueresults = xueresults.replace('雪、', '到')
                xueresults = xueresults.replace('中到小雪', '小到中雪')
                xueresults = xueresults.replace('大到中雪', '中到大雪')

        mincell = grouped_color[line_name + '_mincell']
        maxcell = grouped_color[line_name + '_maxcell']

        minheadcellTime = grouped_color[line_name + '_minheadcellvalue']
        maxheadcellTime = grouped_color[line_name + '_maxheadcellvalue']
        # minheadcellTime = (minheadcellTime - 3 if minheadcellTime - 3 != -1 else 23)

        # 获取最左边和最右边单元格所属表头的日期和时间
        mincellDate = getDate(ws_sheet1, mincell)
        maxcellDate = getDate(ws_sheet1, maxcell)
        mincellDateStr = getDateStr(mincellDate)
        maxcellDateStr = getDateStr(maxcellDate)

        # 解决23时往前推3个小时后，日期变成了前一天的问题
        if minheadcellTime - 3 == -1:
            minheadcellTime = 23
            precell = getprecell(ws_sheet1, mincell)
            if precell == mincell:
                minheadcellTime = 0
            else:
                mincell = precell
                mincellDate = getDate(ws_sheet1, mincell)
                mincellDateStr = getDateStr(mincellDate)
        else:
            minheadcellTime = minheadcellTime - 3

        # ，其中：麻田镇站在25日20时到26日02时有短时大雪或雨夹雪。
        if mincellDate == maxcellDate:
            resultword = f'{line_name}线：在{mincellDateStr}{minheadcellTime:02d}时到{maxheadcellTime:02d}时有{xueresults}'
        else:
            resultword = f'{line_name}线：在{mincellDateStr}{minheadcellTime:02d}时到{maxcellDateStr}{maxheadcellTime:02d}时有{xueresults}'
        result[line_name].append(resultword)
    # print(f'总的 result-->{result}')

    # 【其中】文字部分
    # 如果没有【大雨夹雪】，那么就考虑【大雪】的情况
    for row in ws_sheet1.iter_rows(min_row=4):  # 从数据行开始
        line_name = row[line_col - 1].value  # 线路名
        for cell in row:
            # 获取当前单元格的列索引
            col_index = cell.column
            # 获取每个时间段的单元格颜色
            cell_color = get_colour(cell)
            # 大雪
            if len(grouped_dayujiaxue_cell[line_name]) == 0:
                if cell.column >= 5 and cell_color == '00D945':
                    grouped_daxue_cell[line_name].append(cell)
                # print()

    # print(f'大雪 grouped_daxue_cell-->{grouped_daxue_cell}')
    result = getQizhongAll(ws_sheet1, grouped_daxue_cell, result, 'daxue')
    # print(f'大雨夹雪 grouped_daxue_cell-->{grouped_daxue_cell}')
    result = getQizhongAll(ws_sheet1, grouped_dayujiaxue_cell, result, 'dayujiaxue')

    # print(f'加上【其中】后的result-->{result}')
    # 综合处理【总】和【分】的文字
    finalresult = ''
    for key, value in result.items():
        valueresult = ''
        if len(value) == 1 or len(value[1]) == 0:
            if value[0].find('大雪') == -1:
                continue
            valueresult = value[0]
        else:
            valueresult = '，'.join(value)
            text_to_insert = "其中："
            # 分割字符串
            parts = valueresult.split('，', 1)  # 1 表示只分割一次
            if len(parts) > 1:
                parts[1] = text_to_insert + parts[1]  # 在第二部分前插入文字
            # 合并字符串
            valueresult = '，'.join(parts)
            # print(f'valueresult-->{valueresult}')
            valueresult = valueresult[:-1]
        valueresult += '。'
        # print(f'valueresult-->{valueresult}')
        finalresult += valueresult
        finalresult += '\n'
    # print(f'finalresult-->{finalresult}')

    # 把结果输出到第2行里
    printResultToSecondRow(wb, ws_sheet1, finalresult)

    # 每条【线路名】加上一个超链接，跳转到对应线路的【第一行第一列的单元格】


    return finalresult

def getQizhongAll(ws_sheet1, grouped_dayujiaxue_cell, result, flag):
    grouped_qizhong_data = defaultdict(list)
    # 处理【其中】集合的数据
    for line_name, dayujiaxuelist in grouped_dayujiaxue_cell.items():
        line_list = list()
        sublist_1 = list()
        sublist_2 = list()
        # print(f'dayujiaxuelist-->{dayujiaxuelist}')
        for i in range(1, len(dayujiaxuelist)):
            if (dayujiaxuelist[i].row - dayujiaxuelist[i - 1].row in (0, 1)):
                sublist_1.append(dayujiaxuelist[i - 1])
                sublist_1.append(dayujiaxuelist[i])
            else:
                sublist_2.append(dayujiaxuelist[i - 1])
                sublist_2.append(dayujiaxuelist[i])
        # print('sublist_1', sublist_1)
        # print('sublist_2', sublist_2)
        line_list.append(sublist_1)
        line_list.append(sublist_2)
        # print('line_list', line_list)
        grouped_qizhong_data[line_name] = line_list
    # print(f'grouped_qizhong_data-->{grouped_qizhong_data}')

    qizhongresult = ''
    for line_name, qizhong_data in grouped_qizhong_data.items():
        list1 = qizhong_data[0]
        list2 = qizhong_data[1]
        # print(f'list1-->{list1}')
        # print(f'list2-->{list2}')
        dayujiaxuelist = grouped_dayujiaxue_cell[line_name]  # 全部红色cell：大雨夹雪
        # print(f'dayujiaxuelist-->{dayujiaxuelist}')
        # print(f'line_name-->{line_name}')
        if len(list2) == 0:
            if len(list1) == 0:
                # 没有【大雨夹雪】的情况，eg:唐遵线
                qizhongresult = ''
            else:
                # 【大雨夹雪】没有分段的情况，eg:阳涉线
                startcell = list1[0]
                endcell = list1[-1]
                # print(f'line_name-->{line_name}')
                # print(f'startcell-->', startcell)
                # print(f'endcell-->', endcell)
                # 找到最小列的cell和最大列的cell
                startcell_col, endcell_col = getminmaxcell_col(line_name, dayujiaxuelist, startcell, endcell)
                qizhongresult = getqizhongresult(ws_sheet1, startcell, endcell, startcell_col, endcell_col, flag)
            # print()
        else:
            trunclist = list()
            for i in range(len(list2) - 1):
                if i % 2 == 0:
                    cuplelist = list()
                    cuplelist.append(list2[i])
                    cuplelist.append(list2[i + 1])
                    trunclist.append(cuplelist)
            # print(f'trunclist-->{trunclist}')

            # 第一个片段
            startcell = dayujiaxuelist[0]
            endcell = trunclist[0][0]  # 也就是list2[0]
            # 找到最小列的cell和最大列的cell
            # print(f'startcell-->', startcell)
            # print(f'endcell-->', endcell)
            startcell_col, endcell_col = getminmaxcell_col(line_name, dayujiaxuelist, startcell, endcell)
            qizhongresult = getqizhongresult(ws_sheet1, startcell, endcell, startcell_col, endcell_col, flag)

            # 中间的片段
            for i in range(0, len(trunclist) - 1):
                startcell = trunclist[i][1]
                endcell = trunclist[i + 1][0]
                # print(f'startcell-->', startcell)
                # print(f'endcell-->', endcell)
                # 找到最小列的cell和最大列的cell
                startcell_col, endcell_col = getminmaxcell_col(line_name, dayujiaxuelist, startcell, endcell)
                qizhongresult += getqizhongresult(ws_sheet1, startcell, endcell, startcell_col, endcell_col, flag)

            # 最后一个片段
            startcell = trunclist[-1][1]
            endcell = dayujiaxuelist[-1]
            # print(f'startcell-->{startcell}')
            # print(f'endcell-->{endcell}')
            # 找到最小列的cell和最大列的cell
            startcell_col, endcell_col = getminmaxcell_col(line_name, dayujiaxuelist, startcell, endcell)
            qizhongresult += getqizhongresult(ws_sheet1, startcell, endcell, startcell_col, endcell_col, flag)
            # print()
        if qizhongresult != '':
            result[line_name].append(qizhongresult)
    return result

def getminmaxcell_col(line_name, dayujiaxuelist, startcell, endcell):
    # print('startcell的index是:', line_name, dayujiaxuelist.index(startcell))
    # print('endcell的index是:', line_name, dayujiaxuelist.index(endcell))
    newdayujiaxuelist = dayujiaxuelist[dayujiaxuelist.index(startcell) : dayujiaxuelist.index(endcell) + 1]
    # print('新的集合：', newdayujiaxuelist)
    grouped_color = defaultdict(int)
    for cell in newdayujiaxuelist:
        if grouped_color[line_name + '_mincol'] == 0:
            grouped_color[line_name + '_mincol'] = cell.column
            grouped_color[line_name + '_mincell'] = cell
        elif cell.column < grouped_color[line_name + '_mincol']:
            grouped_color[line_name + '_mincol'] = cell.column
            grouped_color[line_name + '_mincell'] = cell

        if grouped_color[line_name + '_maxcol'] == 0:
            grouped_color[line_name + '_maxcol'] = cell.column
            grouped_color[line_name + '_maxcell'] = cell
        elif cell.column > grouped_color[line_name + '_maxcol']:
            grouped_color[line_name + '_maxcol'] = cell.column
            grouped_color[line_name + '_maxcell'] = cell
    startcell, endcell = grouped_color[line_name + '_mincell'], grouped_color[line_name + '_maxcell']
    # print('startcell, endcell', startcell, endcell)
    # print()
    return startcell, endcell
def getqizhongresult(ws_sheet1, startcell, endcell, startcell_col, endcell_col, flag):
    # 获得车站
    startcellstation = ws_sheet1.cell(row=startcell.row, column=2).value
    endcellstation = ws_sheet1.cell(row=endcell.row, column=2).value
    # print(f'startcellstation-->{startcellstation}')
    # print(f'endcellstation-->{endcellstation}')

    # print(f'startcell_col-->{startcell_col}')
    # print(f'endcell_col-->{endcell_col}')
    # print()
    # 获得日期
    mincellDate = getDate(ws_sheet1, startcell_col)
    maxcellDate = getDate(ws_sheet1, endcell_col)
    mincellDateStr = getDateStr(mincellDate)
    maxcellDateStr = getDateStr(maxcellDate)
    # 获得时间
    mincelltime = ws_sheet1.cell(row=3, column=startcell_col.column).value
    maxcelltime = ws_sheet1.cell(row=3, column=endcell_col.column).value

    # mincelltimetrue = (mincelltime - 3 if mincelltime - 3 != -1 else 23)
    # 解决23时往前推3个小时后，日期变成了前一天的问题
    if mincelltime - 3 == -1:
        mincelltimetrue = 23
        precell = getprecell(ws_sheet1, startcell_col)
        if precell == startcell_col:
            mincelltimetrue = 0
        else:
            startcell_col = precell
            mincellDate = getDate(ws_sheet1, startcell_col)
            mincellDateStr = getDateStr(mincellDate)
    else:
        mincelltimetrue = mincelltime - 3

    # ，其中：麻田镇站在25日20时到26日02时有短时大雪或雨夹雪。
    tianqistr = '或雨夹雪' if flag == 'dayujiaxue' else ''
    if mincellDate == maxcellDate:
        if startcell.row == endcell.row:
            qizhongresult = f'{startcellstation}站在{mincellDateStr}{mincelltimetrue:02d}时到{maxcelltime:02d}时有短时大雪{tianqistr}，'
        else:
            qizhongresult = f'{startcellstation}到{endcellstation}段在{mincellDateStr}{mincelltimetrue:02d}时到{maxcelltime:02d}时有短时大雪{tianqistr}，'
    else:
        if startcell.row == endcell.row:
            qizhongresult = f'{startcellstation}站在{mincellDateStr}{mincelltimetrue:02d}时到{maxcellDateStr}{maxcelltime:02d}时有短时大雪{tianqistr}，'
        else:
            qizhongresult = f'{startcellstation}到{endcellstation}段在{mincellDateStr}{mincelltimetrue:02d}时到{maxcellDateStr}{maxcelltime:02d}时有短时大雪{tianqistr}，'

    # print(f'qizhongresult-->{qizhongresult}')
    return qizhongresult

def getprecell(ws, cell):
    # 获取当前单元格的坐标
    current_coord = cell.coordinate
    # print(current_coord)
    # print(type(current_coord))

    col_num = ord(current_coord[0]) - ord('A') + 1  # 将列字母转换为数字（例如，'B' -> 2）
    # print(f'aa-->{current_coord[0]}')
    # print(f'col_num-->{col_num}')

    # 计算上一列的列号（数字形式）
    prev_col_num = col_num - 1
    # print(f'prev_col_num-->{prev_col_num}')
    if prev_col_num > 0:  # 确保不是第一列（A列）
        # 将列号转换回字母形式
        prev_col_letter = get_column_letter(prev_col_num)
        # print(f'prev_col_letter-->{prev_col_letter}')
        # 构建新的单元格坐标
        prev_cell_coord = f"{prev_col_letter}{cell.row}"
        # print(f'prev_cell_coord-->{prev_cell_coord}')
        # 获取同一行的上一列的单元格
        prev_cell = ws[prev_cell_coord]
        # print(f'prev_cell-->{prev_cell}')
        # print(f"The cell to the left of {prev_cell} is {prev_cell.value}")
        return prev_cell
    else:
        # print("The cell is in the first column (A).")
        return cell
def getDateStr(cellDate):
    return f'{int(cellDate.split('-')[0])}月{cellDate.split('-')[1]}日'
def getDate(sheet, cell):
    # 获取 N4 单元格的值
    row_now = cell.row  # 获取 N4 所在的行号

    # 获取上一行（N3）和再上一行（N2）的值
    cell_row2 = sheet.cell(row=3, column=cell.column)  # N列是第14列
    cell_row1 = sheet.cell(row=2, column=cell.column)

    # 判断 N2 是否是合并单元格
    cell_row1_value = ''
    is_merged = False
    for merged_range in sheet.merged_cells.ranges:
        if sheet.cell(row=2, column=cell.column).coordinate in merged_range:
            is_merged = True
            # 获取合并区域的左上角单元格的值
            left_top_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            cell_row1_value = left_top_cell.value
            # print(f"N2 (合并单元格) 的值：{cell_row1_value}")
            break

    if not is_merged:
        cell_row1_value = cell_row1.value
        # print(f"N2: {cell_row1_value}")
    # 打印 N4 和 N3 的值
    # print(f"N4: {cell.value}")
    # print(f"N3: {cell_row2.value}")
    return cell_row1_value

def printResultToSecondRow(wb, ws_sheet1, finalresult):
    # 获取第一行的合并单元格信息
    merged_cells = list(ws_sheet1.merged_cells.ranges)  # 返回的是一个 MergeCellRange 的列表
    print(merged_cells)
    # 一大段文字，包含换行符
    # text = "阳涉线：在1月25日11时到1月26日02时有中到大雪，其中：麻田镇站在1月25日20时到1月26日02时有短时大雪或雨夹雪。\n唐遵线：在1月25日23时到1月26日20时有小雪，其中：贾庵子到遵化南段在1月26日05时到11时有短时大雪。\n唐曹线：在1月26日02时到20时有小雪，其中：七道桥到曹妃甸东段在1月26日05时到20时有短时大雪。"
    text = finalresult
    # 计算合并的列数并在第二行进行合并
    for merged_range in merged_cells:
        # 只处理第一行的合并区域
        if merged_range.min_row == 1:
            # 获取合并区域的开始列和结束列
            start_col = merged_range.min_col
            end_col = merged_range.max_col
            # 合并第二行的相应单元格
            ws_sheet1.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            # 将文字写入第二行的合并单元格中
            cell = ws_sheet1.cell(row=2, column=start_col)
            cell.value = text
            # 设置字体大小为11（和孙老师商议决定的，但是存在小屏幕显示不全的问题）
            cell.font = Font(size=11)
            # 设置单元格内容换行
            cell.alignment = Alignment(wrap_text=True)
    # 设置行高
    ws_sheet1.row_dimensions[2].height = (text.count('\n')+1) * 15
    # ws_sheet1.row_dimensions[2].height = 500
    # 保存文件
    wb.save('字体11号.xlsx')


# 使用示例
if __name__ == "__main__":
    # file_path = "北京3.xlsx"
    # file_path = "北京1.xlsx"
    # file_path = "北京5.xlsx"
    file_path = "北京局普速精细化预报服务表-2025-01-24.xlsx"

    output = process_weather_data(file_path)
    print(f'output-->{output}')
